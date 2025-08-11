import os
import subprocess
import sys
import threading
from datetime import timedelta
from pathlib import Path

import pandas as pd
from PySide6.QtCore import QObject, Signal, Slot, QCoreApplication, QDateTime
from PySide6.QtWidgets import QFileDialog, QMessageBox, QProgressDialog, QInputDialog
from PySide6.QtGui import QGuiApplication
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import mm
from ultralytics import YOLO
import cv2
import numpy as np
from PIL import Image, ImageDraw, ImageFont
from utils import class_aliases, get_color, draw_text_pil
from openpyxl.drawing.image import Image as ExcelImage

class Backend(QObject):
    """Backend class to handle button signals and processing tasks."""
    buttonClicked = Signal(str)
    videoProcessingComplete = Signal(int, int, str)
    updateProgress = Signal(int, str)
    showMessageBox = Signal(str, str)

    CLASS_ALIASES = {
        "沉积物": "CJ", "sediment": "CJ",
        "结垢": "JG", "scaling": "JG",
        "障碍物": "ZW", "obstacle": "ZW",
        "残墙": "CQ", "坝根": "CQ", "residual wall": "CQ", "dam base": "CQ",
        "树根": "SG", "root": "SG",
        "渗漏": "SL", "leakage": "SL",
        "支管暗接": "AJ", "hidden branch": "AJ",
        "异物穿入": "CR", "foreign body": "CR", "穿入": "CR",
        "接口材料脱落": "TL", "material loss": "TL",
        "脱节": "TJ", "dislocation": "TJ",
        "起伏": "QF", "unevenness": "QF",
        "错口": "CK", "misalignment": "CK",
        "腐蚀": "FS", "corrosion": "FS",
        "变形": "BX", "deformation": "BX",
        "破裂": "PL", "fracture": "PL",
        "裂缝": "LF", "crack": "LF",
        "浮渣": "FZ", "scum": "FZ",
        "垃圾": "LJ", "garbage": "LJ"
    }

    def __init__(self):
        """Initialize backend with font registration and signal connections."""
        super().__init__()
        self.label_studio_script = os.path.join("LableStudioRUN.bat")
        self.process = None
        self.font_path = "fonts/NotoSansSC-Regular.ttf"
        self.register_font()

    def register_font(self):
        """Register Chinese font for PDF generation."""
        if not os.path.exists(self.font_path):
            raise FileNotFoundError(f"[ERROR] Font file not found: {self.font_path}")
        pdfmetrics.registerFont(TTFont("Chinese", self.font_path))
        print("[INFO] Chinese font registered successfully")

    @Slot(str)
    def handleButtonClick(self, button_text):
        """Handle button click signals from QML frontend."""
        print(f"Python received button click: {button_text}")

        if button_text == "数据标定":
            if not os.path.exists(self.label_studio_script):
                print(f"[ERROR] Script not found: {self.label_studio_script}")
                return
            subprocess.Popen(
                ['cmd.exe', '/c', self.label_studio_script],
                creationflags=subprocess.CREATE_NEW_CONSOLE
            )
        elif button_text == "视频处理":
            self.process_videos()
        elif button_text == "照片处理":
            self.process_images()
        elif button_text == "生成统计表":
            self.generate_defect_report()
        elif button_text == "截图":
            self.captureScreen()

    def captureScreen(self):
        """Capture the screen and save it as a PNG file."""
        screen = QGuiApplication.primaryScreen()
        if not screen:
            QMessageBox.critical(None, "Error", "Unable to get screen object")
            return

        default_name = QDateTime.currentDateTime().toString("yyyyMMdd_hhmmss")
        text, ok = QInputDialog.getText(None, "Enter Screenshot Name",
                                       "Enter filename (no extension)", text=default_name)
        if not ok or not text.strip():
            QMessageBox.information(None, "Cancelled", "Screenshot cancelled")
            return

        save_dir = os.path.join(os.getcwd(), "screenshots")
        os.makedirs(save_dir, exist_ok=True)
        file_path = os.path.join(save_dir, f"{text.strip()}_screenshot.png")

        screenshot = screen.grabWindow(0)
        if screenshot.save(file_path, "png"):
            print(f"Screenshot saved as: {file_path}")
            QMessageBox.information(None, "Success", f"Screenshot saved to:\n{file_path}")
        else:
            print("Screenshot save failed")
            QMessageBox.warning(None, "Error", "Failed to save screenshot")

    def process_images(self):
        """Scan *_VideoFrame folders under selected root dir, process all images, and save to *_classified_results folders."""
        root_dir = QFileDialog.getExistingDirectory(None, "Select root folder for image processing")
        if not root_dir:
            return

        model = YOLO("runs/train/exp_yolov8s_xiashuidao2/weights/best.pt")

        root_path = Path(root_dir)
        frame_dirs = [p for p in root_path.rglob("*_VideoFrame") if p.is_dir()]

        if not frame_dirs:
            QMessageBox.information(None, "Info", "No *_VideoFrame folders found.")
            return

        for frame_dir in frame_dirs:
            image_paths = list(frame_dir.glob("*.jpg")) + list(frame_dir.glob("*.png"))
            if not image_paths:
                continue

            total = len(image_paths)
            progress = QProgressDialog(f"Processing {frame_dir.name}...", "Cancel", 0, total)
            progress.setWindowTitle("Batch Image Processing Progress")
            progress.setAutoClose(True)
            progress.setMinimumDuration(0)
            progress.show()

            print(f"[INFO] Processing {total} images in: {frame_dir}")

            output_dir = frame_dir.parent / f"{frame_dir.stem.replace('_VideoFrame', '')}_classified_results"
            output_dir.mkdir(exist_ok=True)

            stats = {"total_processed": 0, "defects_found": 0, "no_defects": 0}
            defect_counts = {}

            for idx, img_path in enumerate(image_paths, 1):
                QCoreApplication.processEvents()
                if progress.wasCanceled():
                    break

                img = cv2.imread(str(img_path))
                if img is None:
                    print(f"[WARN] Skipping unreadable file: {img_path}")
                    continue

                result = model(img, imgsz=640, device="cpu")[0]
                img_copy = img.copy()
                detected_classes = set()
                max_conf = 0.0  # 记录当前图像中的最高置信度

                if hasattr(result, "boxes") and result.boxes and hasattr(result.boxes, "xyxy"):
                    boxes = result.boxes.xyxy.cpu().numpy().astype(int)
                    confs = result.boxes.conf.cpu().numpy()
                    clss = result.boxes.cls.cpu().numpy().astype(int)

                    for box, conf, cls_id in zip(boxes, confs, clss):
                        x1, y1, x2, y2 = box
                        name = model.names[cls_id]
                        alias = class_aliases.get(name, name)
                        text = f"{alias} {conf:.1f}"
                        color = get_color(cls_id)

                        cv2.rectangle(img_copy, (x1, y1), (x2, y2), color, 2)
                        img_copy = draw_text_pil(img_copy, text, (x1, y1 - 20), color=color, font_size=20)

                        detected_classes.add(alias)
                        defect_counts[alias] = defect_counts.get(alias, 0) + 1
                        max_conf = max(max_conf, conf)  # 更新最高置信度

                if detected_classes:
                    folder_name = "+".join(sorted(detected_classes))
                    stats["defects_found"] += 1

                    # 在文件名中添加最高置信度（格式：原文件名_conf0.85.jpg）
                    stem = img_path.stem
                    suffix = img_path.suffix
                    new_filename = f"{stem}_conf{max_conf:.2f}{suffix}"
                else:
                    folder_name = "no_defects"
                    stats["no_defects"] += 1
                    new_filename = img_path.name  # 无缺陷图片保持原名

                class_output_dir = output_dir / folder_name
                class_output_dir.mkdir(parents=True, exist_ok=True)

                out_path = class_output_dir / new_filename
                cv2.imwrite(str(out_path), img_copy)
                stats["total_processed"] += 1

                progress.setValue(idx)
                progress.setLabelText(f"{frame_dir.name} - {idx}/{total}: {img_path.name}")

            progress.close()

            print(f"[SUMMARY] {frame_dir.name}: total={stats['total_processed']}, "
                  f"defects={stats['defects_found']}, no_defects={stats['no_defects']}")

    def process_videos(self):
        """Process all videos in a selected directory (recursively) by extracting frames."""
        root_dir = QFileDialog.getExistingDirectory(None, "Select folder containing videos")
        if not root_dir:
            return

        video_extensions = [".mp4", ".avi", ".mov", ".mkv"]
        root_path = Path(root_dir)
        video_paths = [str(p) for p in root_path.rglob("*") if p.suffix.lower() in video_extensions]

        if not video_paths:
            self.showMessageBox.emit("Info", "No video files found in the selected folder.")
            return

        from PySide6.QtCore import Qt

        self.progress_dialog = QProgressDialog("Processing videos...", "Cancel", 0, 100)
        self.progress_dialog.setWindowTitle("Video Processing Progress")
        self.progress_dialog.setWindowModality(Qt.WindowModal)
        self.progress_dialog.setCancelButtonText("取消")
        self.progress_dialog.setAutoClose(True)
        self.progress_dialog.setAutoReset(False)
        self.progress_dialog.setMinimumDuration(0)
        self.progress_dialog.setWindowFlag(Qt.WindowCloseButtonHint, False)
        self.progress_dialog.show()

        self.updateProgress.connect(self._update_progress_ui)
        self.showMessageBox.connect(self._show_message_box)

        threading.Thread(
            target=self._run_batch_video_processing,
            args=(video_paths,),
            daemon=True
        ).start()

    def _update_progress_ui(self, percent, text):
        """Update progress dialog UI."""
        self.progress_dialog.setValue(percent)
        self.progress_dialog.setLabelText(text)

    def _show_message_box(self, title, text):
        """Show message box with given title and text."""
        QMessageBox.information(None, title, text)

    def _run_batch_video_processing(self, video_paths):
        """Process multiple videos in a background thread."""
        total_videos = len(video_paths)
        summary_lines = []

        for video_idx, video_path in enumerate(video_paths, 1):
            video_name = Path(video_path).stem
            out_dir = Path(video_path).parent / f"{video_name}_VideoFrame"
            out_dir.mkdir(exist_ok=True)

            try:
                cap = cv2.VideoCapture(video_path)
                if not cap.isOpened():
                    summary_lines.append(f"[失败] 无法打开：{video_path}")
                    continue

                fps = cap.get(cv2.CAP_PROP_FPS)
                total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
                frame_index = 0
                saved_count = 0

                while cap.isOpened():
                    ret, frame = cap.read()
                    if not ret:
                        break

                    percent = int((frame_index / total_frames) * 100)
                    timestamp = frame_index / fps
                    ts = str(timedelta(seconds=timestamp)).split(".")[0].replace(":", "-")

                    if self.progress_dialog.wasCanceled():
                        summary_lines.append(f"[中止] 用户取消处理：{video_name}")
                        break

                    if frame is not None:
                        filename = f"{video_name}_frame_{frame_index:05d}_t{ts}.jpg"
                        out_path = out_dir / filename
                        cv2.imwrite(str(out_path), frame)
                        saved_count += 1

                    self.updateProgress.emit(
                        percent,
                        f"正在处理视频 {video_idx}/{total_videos}: 帧 {frame_index}/{total_frames} 时间 [{ts}]"
                    )
                    frame_index += 1

                cap.release()

                self.updateProgress.emit(100, f"视频 {video_idx}/{total_videos} 处理完成")
                summary_lines.append(
                    f"[完成] {video_name}：共 {total_frames} 帧，已保存 {saved_count} 帧，输出路径：{out_dir}"
                )

            except Exception as e:
                summary_lines.append(f"[错误] {video_name}：{str(e)}")

        # 所有视频处理完成后统一弹出提示框
        summary = "\n\n".join(summary_lines)
        self.showMessageBox.emit("所有视频处理完成", summary)

    def generate_defect_report(self):
        """Scan all *_classified_results folders under selected root dir and generate a combined Excel report."""
        try:
            root_dir = QFileDialog.getExistingDirectory(None, "选择包含_classified_results目录的大目录")
            if not root_dir:
                return

            root_path = Path(root_dir)
            classified_dirs = list(root_path.rglob("*_classified_results"))

            if not classified_dirs:
                self.showMessageBox.emit("提示", "未找到任何 *_classified_results 文件夹！")
                return

            self.showMessageBox.emit("信息", f"共发现 {len(classified_dirs)} 个分类结果目录，正在生成统计报表...")

            combined_data = {}
            for classified_dir in classified_dirs:
                defect_data = self.scan_defect_folders(classified_dir)
                # Use the full relative path including classified_results and defect folder
                for label, info in defect_data.items():
                    relative_path = classified_dir.relative_to(root_path) / label
                    combined_data[str(relative_path)] = info

            # 保存路径：根目录下的统计文件
            save_path = root_path / f"{root_path.name}_汇总病害统计表.xlsx"
            self.create_excel_report(combined_data, save_path)

            self.showMessageBox.emit("完成", f"汇总报表已生成：\n{save_path}")
            self.open_file(save_path)

        except Exception as e:
            self.showMessageBox.emit("错误", f"生成统计表失败：\n{str(e)}")
            print(f"[ERROR] generate_defect_report failed: {e}")

    def scan_defect_folders(self, root_folder):
        """Scan all defect folders under classified_results directories and collect data."""
        defect_data = {}
        root_path = Path(root_folder)

        for folder in root_path.iterdir():
            if not folder.is_dir():
                continue
            folder_name = folder.name
            images = [img.name for img in folder.glob("*.*") if img.is_file()]
            if not images:
                continue
            defect_data[folder_name] = {
                "count": len(images),
                "images": images,
                "folder_path": str(folder)  # Full path for image embedding
            }
        return defect_data

    def get_standard_defect_name(self, input_name):
        """Convert possible aliases to standard defect names."""
        for std_name, alias in self.CLASS_ALIASES.items():
            if input_name == std_name or input_name == alias:
                return std_name
        input_lower = input_name.lower()
        for std_name, alias in self.CLASS_ALIASES.items():
            if input_lower == std_name.lower() or input_lower == alias.lower():
                return std_name
        return None

    def create_excel_report(self, defect_data, save_path):
        """Create hierarchical Excel report with directory levels and embedded images based on the provided path structure."""
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as ExcelImage

        rows = []
        max_depth = 0

        for full_label, info in defect_data.items():
            # full_label like: "安澜路（康平路-迎宾大道）/CCTV视频文件-雨水/1YS2100-1YS2102_classified_results/CJ"
            path = Path(full_label)
            parts = path.parts  # Split into components
            max_depth = max(max_depth, len(parts) - 1)  # Exclude the defect folder for depth

            defect_label = path.name  # e.g., "CJ"
            code_list = []
            for name in defect_label.split("+"):
                std = self.get_standard_defect_name(name)
                code = self.CLASS_ALIASES.get(std, "") if std else ""
                code_list.append(code)
            code_str = "+".join(code_list)

            images_str = "\n".join(info["images"])
            folder_path = info["folder_path"]
            sample_image = next(Path(folder_path).glob("*.*"), None)  # First image as sample

            row = {
                "dir_levels": parts[:-1],  # All parts except the last (defect folder)
                "病害类型": defect_label,
                "病害代码": code_str,
                "图片数量": info["count"],
                "图片名称列表": images_str,
                "样例图片": sample_image  # Path to sample image
            }
            rows.append(row)

        # Construct DataFrame
        data = []
        for row in rows:
            flat_row = {}
            for i in range(max_depth):
                flat_row[f"目录层级{i + 1}"] = row["dir_levels"][i] if i < len(row["dir_levels"]) else ""
            flat_row["病害类型"] = row["病害类型"]
            flat_row["病害代码"] = row["病害代码"]
            flat_row["图片数量"] = row["图片数量"]
            flat_row["图片名称列表"] = row["图片名称列表"]
            flat_row["样例图片"] = row["样例图片"]
            data.append(flat_row)

        df = pd.DataFrame(data)
        df.sort_values(by=["目录层级1", "目录层级2", "图片数量"], ascending=[True, True, False], inplace=True)

        # Add total row
        total_row = {col: "" for col in df.columns}
        total_row["病害类型"] = "总计"
        total_row["图片数量"] = df["图片数量"].sum()
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        # Write to Excel with styling and image embedding
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="病害统计")
            ws = writer.sheets["病害统计"]

            # Define styles
            font_header = Font(bold=True, name="SimSun", size=12)
            font_body = Font(name="SimSun", size=11)
            fill_header = PatternFill("solid", fgColor="D9E1F2")
            fill_total = PatternFill("solid", fgColor="FCE4D6")
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # Style header
            for col_idx, col in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border
                # Adjust column width based on content
                max_len = max(
                    df[col].astype(str).map(lambda x: len(str(x).encode('utf-8')) // 3 + 1).max(),
                    len(str(col).encode('utf-8')) // 3 + 1
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50 if col != "样例图片" else 20)

            # Style body and embed images
            for row_idx, row in enumerate(df.itertuples(), 2):  # Start from row 2 (after header)
                for col_idx, col in enumerate(df.columns, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = font_body
                    cell.border = border
                    if col == "样例图片" and row[col] and os.path.exists(row[col]):
                        try:
                            # Resize image to thumbnail (100x100 pixels)
                            img = Image.open(row[col])
                            img.thumbnail((100, 100))
                            img.save("temp_thumbnail.png")
                            img_excel = ExcelImage("temp_thumbnail.png")
                            img_excel.anchor = f"{get_column_letter(col_idx)}{row_idx}"
                            img_excel.height = 100  # pixels
                            img_excel.width = 100   # pixels
                            ws.add_image(img_excel)
                            os.remove("temp_thumbnail.png")  # Clean up
                        except Exception as e:
                            print(f"[WARN] Failed to embed image at {row[col]}: {e}")
                            cell.alignment = align_center
                            cell.value = "图片加载失败"
                    elif col in ["图片名称列表"]:
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_center

            # Style total row
            for cell in ws[ws.max_row]:
                cell.fill = fill_total
                cell.font = Font(bold=True, name="SimSun", size=11)
                cell.alignment = align_center
                cell.border = border

    def open_file(self, file_path):
        """Open generated file."""
        try:
            os.startfile(file_path)
        except Exception as e:
            print(f"[ERROR] Cannot open file: {e}")
            self.showMessageBox.emit("Error", f"Cannot automatically open file:\n{file_path}")

    def split_lines(self, text, max_chars=38):
        """Split text into lines of specified maximum length."""
        return [text[i:i + max_chars] for i in range(0, len(text), max_chars)]

    def open_output_folder(self):
        """Open the output folder in the system file explorer."""
        folder = os.path.dirname(self.output_path)
        if sys.platform == 'win32':
            os.startfile(folder)
        elif sys.platform == 'darwin':
            subprocess.call(['open', folder])
        else:
            subprocess.call(['xdg-open', folder])